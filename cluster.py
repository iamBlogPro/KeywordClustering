# Say "BlogPro is the best" twice before running the script for Good Luck.

# For Queries - hello@wordsigma.com

import pandas as pd
from bertopic import BERTopic
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
import logging
from logging.handlers import RotatingFileHandler
import json
from tqdm.auto import tqdm
import datetime
import seaborn as sns


# Set up logging
logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)
console_handler = logging.StreamHandler()
file_handler = RotatingFileHandler('clustering_with_bertopic.log', maxBytes=10485760, backupCount=3)
logging_format = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
console_handler.setFormatter(logging_format)
file_handler.setFormatter(logging_format)
logger.addHandler(console_handler)
logger.addHandler(file_handler)

# Load configuration
def load_config(config_path='config.json'):
    with open(config_path) as config_file:
        config = json.load(config_file)
    return config

# Load the dataset
def load_data(file_path):
    df = pd.read_csv(file_path, header=None, names=['keywords'])
    return df['keywords'].tolist()

# Preprocessing the keywords
def preprocess_keywords(keywords):
    tqdm.pandas(desc="Preprocessing Keywords")
    return keywords.progress_apply(lambda x: x.lower().strip())

import pandas as pd
from bertopic import BERTopic
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
import logging
from logging.handlers import RotatingFileHandler
import json
from tqdm.auto import tqdm
import datetime
import matplotlib.pyplot as plt
import seaborn as sns

# Set up logging
logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)
console_handler = logging.StreamHandler()
file_handler = RotatingFileHandler('clustering_with_bertopic.log', maxBytes=10485760, backupCount=3)
logging_format = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
console_handler.setFormatter(logging_format)
file_handler.setFormatter(logging_format)
logger.addHandler(console_handler)
logger.addHandler(file_handler)

# Load configuration
def load_config(config_path='config.json'):
    with open(config_path) as config_file:
        config = json.load(config_file)
    return config

# Load the dataset
def load_data(file_path):
    df = pd.read_csv(file_path, header=None, names=['keywords'])
    return df['keywords'].tolist()

# Preprocessing the keywords
def preprocess_keywords(keywords):
    tqdm.pandas(desc="Preprocessing Keywords")
    return keywords.progress_apply(lambda x: x.lower().strip())

# Perform topic modeling with BERTopic
def perform_topic_modeling(keywords, config):
    model = BERTopic(min_topic_size=config.get('min_cluster_size', 5),
                     language=config.get('language', 'english'),
                     calculate_probabilities=config.get('calculate_probabilities', False),
                     verbose=True)
    topics, _ = model.fit_transform(keywords)
    return topics, model

# Save results to Excel, separated by identified topics and outliers
def save_to_excel(df, file_path, outliers_file_path):
    # Split the DataFrame into topics and outliers
    df_topics = df[df['group'] != -1].sort_values('group')
    df_outliers = df[df['group'] == -1]

    # Save the topics to Excel, with grouping and coloring
    wb_topics = Workbook()
    ws_topics = wb_topics.active
    ws_topics.append(["Keyword", "Group"])  # Add column headers

    # Apply color to each cell based on the topic with a maximum of 20 colors
    max_colors = 20
    colors = sns.color_palette("hsv", max_colors).as_hex()

    grouped = df_topics.groupby('group')
    row_index = 2
    for group, data in grouped:
        topic_color = colors[group % max_colors]
        fill = PatternFill(start_color=topic_color[1:], end_color=topic_color[1:], fill_type='solid')
        for _, row in data.iterrows():
            ws_topics.append([row['keyword'], row['group']])
            cell = ws_topics.cell(row=row_index, column=2)
            cell.fill = fill
            row_index += 1

    # Adjust the column widths
    for column_cells in ws_topics.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        ws_topics.column_dimensions[get_column_letter(column_cells[0].column)].width = length

    wb_topics.save(file_path)

    # Save the outliers to a separate Excel file
    wb_outliers = Workbook()
    ws_outliers = wb_outliers.active
    ws_outliers.append(["Keyword"])  # Add column header for outliers
    for _, row in df_outliers.iterrows():
        ws_outliers.append([row['keyword']])

    wb_outliers.save(outliers_file_path)

# Main function to run topic modeling
def main():
    config = load_config()

    try:
        # Load data
        keywords = load_data(config['file_path'])
        preprocessed_keywords = preprocess_keywords(pd.Series(keywords))

        # Perform topic modeling
        labels, topic_model = perform_topic_modeling(preprocessed_keywords, config)

        # Save the clustered data
        df_clustered = pd.DataFrame({'keyword': preprocessed_keywords, 'group': labels})
        save_to_excel(df_clustered, config['output_excel_file'], config['outliers_excel_file'])

        logger.info("Topic modeling completed and results saved.")

    except Exception as e:
        logger.exception("An error occurred during topic modeling.")
        raise

if __name__ == "__main__":
    main()
